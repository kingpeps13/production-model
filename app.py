import streamlit as st
import math
import json
import pandas as pd
from datetime import datetime, timedelta, date
import plotly.express as px
import io
from openpyxl import Workbook

st.set_page_config(page_title="Модель расчёта производства", layout="wide")
st.title("🏭 Модель расчёта календарного времени выполнения заказа")

# ================== Инициализация сессии ==================
defaults = {
    "operations": [],
    "grammovki": [],
    "gram_counts": {},
    "product_name": "",
    "shift_start": 8.0,
    "shift_duration": 8.0,
    "is_glue": False,
    "result": None,
    "template_name": "template",
    "correction_choice": False,
    "pn_input": "",
    "ss_input": 8.0,
    "sd_input": 8.0,
    "ig_input": False,
    "gs_input": [],
    "q_input": 1200,
    "n_input": 600,
    "template_name_input": "template",
    "pending_template_content": None,
    "start_date_input": date.today()
}
for key, default in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ================== Шаблоны ==================
def template_to_json():
    data = {
        "product_name": st.session_state.pn_input,
        "shift_start": st.session_state.ss_input,
        "shift_duration": st.session_state.sd_input,
        "is_glue": st.session_state.ig_input,
        "grammovki": st.session_state.gs_input if st.session_state.ig_input else [],
        "gram_counts": {g: st.session_state.get(f"g_{g}", 0) for g in st.session_state.gs_input},
        "operations": st.session_state.operations,
        "start_date": st.session_state.start_date_input.isoformat(),
        "version": "2.3.0"
    }
    return json.dumps(data, ensure_ascii=False, indent=2)

def load_template_from_json(json_str):
    data = json.loads(json_str)
    st.session_state.pn_input = data.get('product_name', "")
    st.session_state.ss_input = data.get('shift_start', 8.0)
    st.session_state.sd_input = data.get('shift_duration', 9.0)
    st.session_state.ig_input = data.get('is_glue', False)
    st.session_state.gs_input = data.get('grammovki', [])
    gram_counts = data.get('gram_counts', {})
    for g in [3, 5, 10]:
        st.session_state[f"g_{g}"] = gram_counts.get(g, 0)
    st.session_state.operations = data.get('operations', [])
    if 'start_date' in data:
        try:
            st.session_state.start_date_input = date.fromisoformat(data['start_date'])
        except:
            pass
    st.session_state.result = None

def clear_all():
    keys = ['pn_input', 'ss_input', 'sd_input', 'ig_input', 'gs_input',
            'q_input', 'n_input', 'operations', 'result', 'correction_choice',
            'pending_template_content', 'start_date_input']
    for k in keys:
        if k in st.session_state: del st.session_state[k]
    st.session_state.pn_input = ""
    st.session_state.ss_input = 8.0
    st.session_state.sd_input = 8.0
    st.session_state.ig_input = False
    st.session_state.gs_input = []
    st.session_state.operations = []
    st.session_state.result = None
    st.session_state.correction_choice = False
    st.session_state.start_date_input = date.today()
    st.rerun()

# ================== НОВАЯ СИМУЛЯЦИЯ (параллельная, с правильным min_batch) ==================
@st.cache_data(ttl=3600, show_spinner=False)
def calculate_cached(product_name, shift_start, shift_duration, operations, is_glue,
                     gram_counts_tuple, Q, N, correction_choice, start_date_iso):
    gram_counts = dict(gram_counts_tuple)
    start_date = date.fromisoformat(start_date_iso)
    return calculate(
        product_name, shift_start, shift_duration, operations,
        is_glue, gram_counts, Q, N, correction_choice, start_date
    )

def calculate(product_name, shift_start, shift_duration, operations, is_glue,
              gram_counts, Q, N, correction_choice, start_date):
    hours_per_day = shift_duration
    base_datetime = datetime.combine(start_date, datetime.min.time()) + timedelta(hours=shift_start)

    # ---- Клей (без изменений) ----
    can_count_4kg = 0; can_count_1kg = 0
    shortage_4kg = 0.0; shortage_1kg = 0.0
    total_weight = 0.0
    weight_map = {3: 3.36, 5: 5.6, 10: 11.2}
    corrected = False

    if is_glue:
        total_weight = sum(cnt * weight_map.get(g, 0) for g, cnt in gram_counts.items())
        can_count_4kg = math.ceil(total_weight / 4000.0)
        rem4 = total_weight % 4000.0
        shortage_4kg = 0.0 if rem4 == 0 else 4000.0 - rem4
        can_count_1kg = math.ceil(total_weight / 1000.0)
        rem1 = total_weight % 1000.0
        shortage_1kg = 0.0 if rem1 == 0 else 1000.0 - rem1
        if rem4 != 0 and correction_choice:
            need_weight = shortage_4kg
            max_g = max(gram_counts.keys(), key=lambda g: weight_map.get(g, 0))
            dose_weight = weight_map[max_g]
            add_doses = math.ceil(need_weight / dose_weight)
            gram_counts[max_g] += add_doses
            total_weight += add_doses * dose_weight
            corrected = True
            Q = sum(gram_counts.values())
            can_count_4kg = math.ceil(total_weight / 4000.0)
            rem4 = total_weight % 4000.0
            shortage_4kg = 0.0 if rem4 == 0 else 4000.0 - rem4
            can_count_1kg = math.ceil(total_weight / 1000.0)
            rem1 = total_weight % 1000.0
            shortage_1kg = 0.0 if rem1 == 0 else 1000.0 - rem1

    # ---- Подготовка операций ----
    for op in operations:
        op.setdefault('daily_setup', False)
        op.setdefault('max_hours_per_day', hours_per_day)
        if op.get('manual', False):
            op['capacity'] = op['prod'] * op['people']
        else:
            op['capacity'] = op['prod'] * op.get('equip', 1)
        op.setdefault('min_batch', 1)

    m = math.ceil(Q / N)
    t_per_job = [N / op['capacity'] for op in operations]

    # Состояния для каждой операции
    eq_free_time = [0.0] * len(operations)          # время освобождения оборудования
    queues = [[] for _ in range(len(operations))]   # очереди ожидающих нарядов (индексы)
    intervals = [[] for _ in range(len(operations))] # для каждой операции храним (start, end)
    all_intervals = []                               # общий список (start_h, end_h, label, color)
    colors = px.colors.qualitative.Plotly * 10

    # Время готовности каждого наряда к операции i (изначально 0 для операции 0)
    # ready[i][j] - время, когда наряд j доступен для операции i
    ready = [ [0.0]*m for _ in range(len(operations)) ]
    # Для операции 0 все наряды готовы в 0
    queues[0] = list(range(m))

    # Количество нарядов, поступивших в очередь операции i (включая уже обработанные)
    total_arrived = [0] * len(operations)
    total_arrived[0] = m

    # Флаг завершения операций
    finished_jobs = [0] * len(operations)   # сколько нарядов уже обработано на операции i

    # Вспомогательная функция размещения наряда с учётом дней и наладок
    def schedule_job(op_idx, job_idx, start_time):
        """Размещает один наряд job_idx на операции op_idx начиная со start_time.
        Возвращает время окончания, обновляет intervals и all_intervals."""
        op = operations[op_idx]
        t_i = t_per_job[op_idx]
        remaining = t_i
        current = start_time
        while remaining > 1e-9:
            day_start = (int(current // 24)) * 24
            day_end = day_start + op.get('max_hours_per_day', hours_per_day)
            # Ежедневная наладка
            if op.get('daily_setup', False):
                # проверяем, есть ли уже наладка в этом дне на этом оборудовании
                setup_done = False
                for s, e in intervals[op_idx]:
                    if s >= day_start and s < day_start + op['setup']:
                        setup_done = True
                        break
                if not setup_done and op['setup'] > 0:
                    setup_start = day_start
                    setup_end = min(day_start + op['setup'], day_end)
                    if setup_end > setup_start:
                        intervals[op_idx].append((setup_start, setup_end))
                        all_intervals.append((setup_start, setup_end,
                                               f"Наладка {op['name']}", 'gray'))
            # Доступное время после уже запланированных работ
            used = 0.0
            for s, e in intervals[op_idx]:
                if s < day_end and e > day_start:
                    used += min(e, day_end) - max(s, day_start)
            available = op.get('max_hours_per_day', hours_per_day) - used
            if available < 1e-9:
                current = (int(current // 24) + 1) * 24
                continue
            chunk = min(remaining, available)
            chunk_start = max(current, day_start)
            if chunk_start < day_start:
                chunk_start = day_start
            chunk_end = chunk_start + chunk
            if chunk_end > day_end:
                chunk_end = day_end
                chunk = chunk_end - chunk_start
            intervals[op_idx].append((chunk_start, chunk_end))
            label = f"{op['name']} (нар.{job_idx+1})"
            all_intervals.append((chunk_start, chunk_end, label, colors[op_idx % len(colors)]))
            remaining -= chunk
            current = chunk_end
            if remaining > 1e-9:
                current = (int(current // 24) + 1) * 24
        return current

    # Основной цикл событий
    t = 0.0
    max_iter = 100000
    iter_count = 0
    progress_bar = st.progress(0, text="Симуляция...") if m > 5 else None

    while finished_jobs[-1] < m and iter_count < max_iter:
        iter_count += 1
        # Ищем операцию, которая может начать обработку
        started = False
        for i, op in enumerate(operations):
            # Условия старта:
            # - оборудование свободно (eq_free_time[i] <= t)
            # - в очереди есть наряды
            if eq_free_time[i] > t or not queues[i]:
                continue
            # Проверяем, достаточно ли нарядов для запуска min_batch
            min_b = op.get('min_batch', 1)
            can_start = False
            if len(queues[i]) >= min_b:
                can_start = True
            # Если все наряды уже поступили (total_arrived[i] == m) и очередь не пуста, можно начинать с любым остатком
            if total_arrived[i] == m and len(queues[i]) > 0:
                can_start = True
            # Для последней операции можно начинать, даже если min_batch не достигнут, но все наряды уже пришли (total_arrived == m)
            if i == len(operations)-1 and total_arrived[i] == m and len(queues[i]) > 0:
                can_start = True

            if not can_start:
                continue

            # Определяем размер batch
            batch_size = min(len(queues[i]), min_b) if can_start and total_arrived[i] != m else len(queues[i])
            # Сортируем ожидающие наряды по времени готовности
            batch_indices = sorted(queues[i][:batch_size], key=lambda j: ready[i][j])
            queues[i] = queues[i][batch_size:]   # убираем из очереди

            # Время начала batch: max(t, max(ready), eq_free_time)
            start_t = max(t, max(ready[i][j] for j in batch_indices), eq_free_time[i])
            # Обрабатываем наряды последовательно
            for j in batch_indices:
                start_t = max(start_t, ready[i][j])
                end_t = schedule_job(i, j, start_t)
                # Обновляем готовность для следующей операции
                if i + 1 < len(operations):
                    ready[i+1][j] = end_t
                    queues[i+1].append(j)
                    total_arrived[i+1] += 1
                start_t = end_t  # следующий наряд начнётся после этого
                finished_jobs[i] += 1
            eq_free_time[i] = end_t
            started = True
            break   # на каждой итерации запускаем только одну операцию (можно поменять, но для простоты)

        if not started:
            # Ни одна операция не может начаться – продвигаем время к ближайшему освобождению или к моменту готовности наряда
            next_t = float('inf')
            for i in range(len(operations)):
                if eq_free_time[i] > t:
                    next_t = min(next_t, eq_free_time[i])
                if queues[i]:
                    # минимальное время готовности ожидающих нарядов
                    min_ready = min(ready[i][j] for j in queues[i])
                    if min_ready > t:
                        next_t = min(next_t, min_ready)
            if next_t == float('inf'):
                break
            t = next_t
        # Обновление прогресс-бара
        if progress_bar and iter_count % 10 == 0:
            progress_bar.progress(min(finished_jobs[-1] / m, 1.0), text=f"Обработано {finished_jobs[-1]}/{m} нарядов")

    if progress_bar:
        progress_bar.empty()

    # Итоговое время (максимальное окончание среди всех интервалов)
    T = max((end for _, end, _, _ in all_intervals), default=0.0)
    days_needed = math.ceil(T / 24)

    # Трудоёмкость и статистика
    days_work_list = []
    setup_total_list = []
    labor_details = []
    total_labor = 0.0

    for i, op in enumerate(operations):
        days_set = set()
        for s, e in intervals[i]:
            days_set.add(int(s // 24))
        days_work = len(days_set)
        days_work_list.append(days_work)

        total_work = m * t_per_job[i]
        if op.get('daily_setup', False):
            setup_total = op['setup'] * days_work
        else:
            setup_total = op['setup']
        setup_total_list.append(setup_total)
        labor_i = op['people'] * (total_work + setup_total)
        total_labor += labor_i
        labor_details.append((op['name'], labor_i))

    t_max = max(t_per_job) if t_per_job else 0
    idx_max = t_per_job.index(t_max) if t_per_job else 0
    bottleneck_name = operations[idx_max]['name'] if operations else ""

    # Загрузка по дням
    day_usage_dict = {}
    for day in range(days_needed):
        day_start = day * 24
        day_end = day_start + hours_per_day
        usage = {}
        for i, op in enumerate(operations):
            total_hours = 0.0
            for s, e in intervals[i]:
                if s < day_end and e > day_start:
                    total_hours += min(e, day_end) - max(s, day_start)
            if total_hours > 0:
                usage[op['name']] = total_hours
        day_usage_dict[day] = usage

    return {
        'Q': Q, 'N': N, 'm': m, 'T': T, 'days_needed': days_needed,
        'total_labor': total_labor,
        'name_list': [op['name'] for op in operations],
        't_list': t_per_job,
        'setup_list': [op['setup'] for op in operations],
        'people_list': [op['people'] for op in operations],
        'daily_setup_list': [op.get('daily_setup', False) for op in operations],
        'max_hours_list': [op.get('max_hours_per_day', hours_per_day) for op in operations],
        'days_work_list': days_work_list,
        'setup_total_list': setup_total_list,
        'labor_details': labor_details,
        'bottleneck_name': bottleneck_name,
        't_max': t_max,
        'all_intervals': all_intervals,
        'day_usage_dict': day_usage_dict,
        'shift_start': shift_start,
        'hours_per_day': hours_per_day,
        'product_name': product_name,
        'operations': operations,
        'is_glue': is_glue,
        'can_count_4kg': can_count_4kg,
        'shortage_4kg': shortage_4kg,
        'can_count_1kg': can_count_1kg,
        'shortage_1kg': shortage_1kg,
        'total_weight': total_weight,
        'gram_counts': gram_counts,
        'corrected': corrected,
        'base_datetime': base_datetime
    }

# ================== ИНТЕРФЕЙС (без изменений) ==================
tab1, tab2, tab3 = st.tabs(["📋 Параметры заказа", "🔧 Операции", "💾 Шаблоны"])

with tab1:
    col1, col2, col3 = st.columns(3)
    with col1:
        st.text_input("Наименование продукта", key='pn_input')
        st.date_input("Дата начала производства", key='start_date_input')
    with col2:
        st.number_input("Начало смены (ч)", min_value=0.0, max_value=23.0, step=0.5, key='ss_input')
        st.checkbox("Это клей?", key='ig_input')
    with col3:
        st.number_input("Длительность смены (ч)", min_value=1.0, max_value=24.0, step=0.5, key='sd_input')
        if st.session_state.ig_input:
            st.subheader("🧴 Граммовки клея")
            all_gram = [3, 5, 10]
            st.multiselect("Выберите граммовки", all_gram, key='gs_input')
            for g in st.session_state.gs_input:
                st.number_input(f"Количество {g} мл", min_value=0, step=50, key=f"g_{g}")
            total_q = sum(st.session_state.get(f"g_{g}", 0) for g in st.session_state.gs_input)
            st.info(f"Общий заказ: {total_q} шт")
            st.checkbox("Корректировать до полных 4-кг канистр", key='correction_choice')
        else:
            st.number_input("Количество штук в заказе", min_value=1, step=100, key='q_input')
            st.session_state.correction_choice = False
        st.number_input("Размер наряда (шт)", min_value=1, step=50, key='n_input')

with tab2:
    st.subheader("Список операций")
    if not st.session_state.operations:
        st.info("Добавьте хотя бы одну операцию")

    hcols = st.columns([2, 1, 1, 1, 1, 1, 1, 0.8])
    with hcols[0]: st.markdown("**Название**")
    with hcols[1]: st.markdown("**Произв-ть (шт/ч)**")
    with hcols[2]: st.markdown("**Наладка (ч)**")
    with hcols[3]: st.markdown("**Оборуд.**")
    with hcols[4]: st.markdown("**Людей**")
    with hcols[5]: st.markdown("**Ежедн. нал.**")
    with hcols[6]: st.markdown("**Ручная**")
    with hcols[7]: st.markdown("**Мин. нар.**")

    for i, op in enumerate(st.session_state.operations):
        cols = st.columns([2, 1, 1, 1, 1, 1, 1, 0.8])
        with cols[0]:
            st.text_input("Название", value=op['name'], key=f"name_{i}", label_visibility="collapsed")
        with cols[1]:
            st.number_input("Произв-ть", min_value=0.1, value=op['prod'], key=f"prod_{i}", label_visibility="collapsed")
        with cols[2]:
            st.number_input("Наладка", min_value=0.0, step=0.05, value=op['setup'], key=f"setup_{i}", label_visibility="collapsed")
        with cols[3]:
            if not st.session_state.get(f"manual_{i}", False):
                st.number_input("Оборуд.", min_value=1, value=op.get('equip', 1), key=f"equip_{i}", label_visibility="collapsed")
            else:
                st.markdown("—")
        with cols[4]:
            st.number_input("Людей", min_value=1, value=op.get('people', 1), key=f"people_{i}", label_visibility="collapsed")
        with cols[5]:
            st.checkbox("Ежедн.", value=op.get('daily_setup', False), key=f"daily_{i}", label_visibility="collapsed")
        with cols[6]:
            is_manual = st.checkbox("Ручная", value=op.get('manual', False), key=f"manual_{i}", label_visibility="collapsed")
            st.session_state.operations[i]['manual'] = is_manual
        with cols[7]:
            st.number_input("Мин.нар.", min_value=1, value=op.get('min_batch', 1), key=f"minbatch_{i}", label_visibility="collapsed")

    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        if st.button("➕ Добавить операцию"):
            new_op = {
                "name": f"Операция {len(st.session_state.operations)+1}",
                "prod": 100.0,
                "setup": 0.0,
                "equip": 1,
                "people": 1,
                "daily_setup": False,
                "max_hours_per_day": st.session_state.sd_input,
                "manual": False,
                "min_batch": 1
            }
            st.session_state.operations.append(new_op)
            st.rerun()
    with btn_col2:
        if st.button("🗑️ Удалить последнюю"):
            if len(st.session_state.operations) > 1:
                st.session_state.operations.pop()
                st.rerun()
            else:
                st.warning("Нельзя удалить последнюю операцию")

with tab3:
    st.subheader("Управление шаблонами")
    uploaded = st.file_uploader("Выберите JSON-шаблон", type=["json"], key="template_uploader")
    if uploaded is not None:
        st.session_state.pending_template_content = uploaded.read().decode('utf-8')
    if st.button("📥 Загрузить шаблон", disabled=st.session_state.pending_template_content is None):
        try:
            load_template_from_json(st.session_state.pending_template_content)
            st.success("Шаблон загружен!")
            st.session_state.pending_template_content = None
            st.rerun()
        except Exception as e:
            st.error(f"Ошибка: {e}")

    st.text_input("Имя шаблона для сохранения", key='template_name_input')
    json_data = template_to_json()
    st.download_button("💾 Скачать шаблон", json_data,
                       file_name=f"{st.session_state.template_name_input or 'template'}.json",
                       mime="application/json")
    st.divider()
    if st.button("🧹 Очистить всё", type="secondary"):
        clear_all()

# Кнопка расчёта
st.divider()
if st.button("🚀 Рассчитать", type="primary", use_container_width=True):
    ops = []
    for i in range(len(st.session_state.operations)):
        op = {
            "name": st.session_state.get(f"name_{i}", ""),
            "prod": st.session_state.get(f"prod_{i}", 0.0),
            "setup": st.session_state.get(f"setup_{i}", 0.0),
            "equip": st.session_state.get(f"equip_{i}", 1),
            "people": st.session_state.get(f"people_{i}", 1),
            "daily_setup": st.session_state.get(f"daily_{i}", False),
            "max_hours_per_day": st.session_state.get(f"maxh_{i}", st.session_state.sd_input),
            "manual": st.session_state.get(f"manual_{i}", False),
            "min_batch": st.session_state.get(f"minbatch_{i}", 1)
        }
        ops.append(op)
    st.session_state.operations = ops

    product_name = st.session_state.pn_input
    shift_start = st.session_state.ss_input
    shift_duration = st.session_state.sd_input
    is_glue = st.session_state.ig_input
    gram_counts_tuple = tuple((g, st.session_state.get(f"g_{g}", 0)) for g in st.session_state.gs_input)
    if is_glue:
        Q = sum(st.session_state.get(f"g_{g}", 0) for g in st.session_state.gs_input)
    else:
        Q = st.session_state.get('q_input', 1200)
    N = st.session_state.get('n_input', 600)
    correction = st.session_state.correction_choice if is_glue else False
    start_date_iso = st.session_state.start_date_input.isoformat()

    with st.spinner("Выполняется расчёт..."):
        result = calculate_cached(
            product_name, shift_start, shift_duration, ops,
            is_glue, gram_counts_tuple, Q, N, correction, start_date_iso
        )
    st.session_state.result = result
    if result.get('corrected'):
        for g, cnt in result['gram_counts'].items():
            st.session_state[f"g_{g}"] = cnt
    st.rerun()

# ================== Отображение результатов ==================
if st.session_state.result is not None:
    result = st.session_state.result
    st.success("✅ Расчёт завершён!")

    with st.expander("🔧 Отладка интервалов"):
        st.write(f"Всего интервалов: {len(result['all_intervals'])}")
        unique_ops = set()
        for _, _, label, _ in result['all_intervals']:
            if not label.startswith("Наладка"):
                op_name = label.split(" (нар.")[0]
            else:
                op_name = label.replace("Наладка ", "")
            unique_ops.add(op_name)
        st.write("Операции в интервалах:", sorted(unique_ops))

    if result['is_glue'] and result['corrected']:
        st.info(f"📝 Заказ скорректирован. Новое количество: {result['Q']} шт. Вес: {result['total_weight']:.2f} г.")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📦 Заказ", f"{result['Q']} шт")
    col2.metric("📋 Нарядов", result['m'])
    col3.metric("⏱️ Календарное время", f"{result['T']:.2f} ч")
    col4.metric("📅 Рабочих дней", result['days_needed'])

    if result['is_glue']:
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("🧴 Общий вес", f"{result['total_weight']:.2f} г")
        c2.metric("📦 4-кг канистр", result['can_count_4kg'])
        if result['shortage_4kg'] > 0:
            c3.metric("⚠️ Недостаток 4-кг", f"{result['shortage_4kg']:.2f} г")
        else:
            c3.metric("✅ 4-кг", "кратно")
        c4.metric("📦 1-кг канистр", result['can_count_1kg'])
        if result['shortage_1kg'] > 0:
            c5.metric("⚠️ Недостаток 1-кг", f"{result['shortage_1kg']:.2f} г")
        else:
            c5.metric("✅ 1-кг", "кратно")

    st.metric("🏭 Узкое место", f"{result['bottleneck_name']} ({result['t_max']:.2f} ч/наряд)")
    st.metric("👷 Общая трудоёмкость", f"{result['total_labor']:.2f} чел·ч")

    st.subheader("📊 Детализация по операциям")
    df_ops = pd.DataFrame({
        "Операция": result['name_list'],
        "t_i (ч)": result['t_list'],
        "Наладка (ч)": result['setup_list'],
        "Людей": result['people_list'],
        "Ежедн. наладка": result['daily_setup_list'],
        "Общее время (ч)": [result['m'] * t for t in result['t_list']],
        "Дней работы": result['days_work_list'],
        "Трудоёмкость (чел·ч)": [lab for _, lab in result['labor_details']]
    })
    st.dataframe(df_ops, use_container_width=True)

    st.subheader("📅 Загрузка по дням")
    if result['day_usage_dict']:
        rows = []
        for day, usage in result['day_usage_dict'].items():
            row = {"День": day + 1}
            row.update(usage)
            rows.append(row)
        df_days = pd.DataFrame(rows)
        if not df_days.empty:
            st.dataframe(df_days, use_container_width=True)
    else:
        st.info("Нет данных по дням")

    # ========== ДИАГРАММА ГАНТА ==========
    st.subheader("📈 Диаграмма Ганта")
    all_intervals_hours = result['all_intervals']
    if all_intervals_hours:
        base_dt = result['base_datetime']
        df_rows = []
        for start_h, end_h, label, color in all_intervals_hours:
            if end_h <= start_h:
                continue
            start_dt = base_dt + timedelta(hours=start_h)
            end_dt = base_dt + timedelta(hours=end_h)
            if label.startswith("Наладка"):
                operation = label.replace("Наладка ", "").strip()
                group = "Наладка"
            else:
                operation = label.split(" (нар.")[0].strip() if " (нар." in label else label.strip()
                group = operation
            df_rows.append({
                "Операция": operation,
                "Начало": start_dt,
                "Окончание": end_dt,
                "Группа": group,
                "Описание": label,
                "Длительность (ч)": end_h - start_h
            })
        df_gantt = pd.DataFrame(df_rows)

        if not df_gantt.empty:
            op_list = result['name_list']
            palette = px.colors.qualitative.Plotly
            color_map = {op: palette[i % len(palette)] for i, op in enumerate(op_list)}
            color_map["Наладка"] = "gray"

            fig = px.timeline(
                df_gantt,
                x_start="Начало",
                x_end="Окончание",
                y="Операция",
                color="Группа",
                color_discrete_map=color_map,
                hover_name="Описание",
                hover_data={
                    "Начало": True,
                    "Окончание": True,
                    "Группа": False,
                    "Длительность (ч)": True,
                    "Описание": False,
                },
                title=f'Диаграмма Ганта для заказа {result["product_name"]} ({result["Q"]} шт)',
                labels={"Операция": "Операция"}
            )

            fig.update_yaxes(
                autorange="reversed",
                categoryorder='array',
                categoryarray=op_list,
                title="Операция"
            )

            fig.update_xaxes(
                title="Дата и время",
                tickformat="%d.%m %H:%M",
                showgrid=True,
                rangeslider_visible=True
            )

            finish_dt = base_dt + timedelta(hours=result['T'])
            fig.add_vline(x=finish_dt, line_width=2, line_dash="dash", line_color="red")
            fig.add_annotation(
                x=finish_dt,
                y=1,
                yref="paper",
                text=f"Конец заказа<br>{result['T']:.2f} ч",
                showarrow=False,
                bgcolor="white",
                font=dict(size=12)
            )

            fig.update_layout(
                height=max(450, len(op_list) * 90),
                hoverlabel=dict(bgcolor="white", font_size=13)
            )

            st.plotly_chart(fig, use_container_width=True)

            total_objs = len(df_gantt)
            ops_objs = sum(1 for _, _, label, _ in all_intervals_hours if not label.startswith("Наладка"))
            st.metric("🧩 Объектов на диаграмме", total_objs)
            st.caption(f"Из них: {ops_objs} операций, {total_objs - ops_objs} наладок")

            with st.expander("🔍 Данные для Ганта"):
                st.dataframe(df_gantt)
        else:
            st.warning("Нет данных для отображения")
    else:
        st.info("Нет данных для построения диаграммы")

    # ================== Экспорт в Excel ==================
    st.subheader("💾 Экспорт")
    try:
        wb = Workbook()
        wb.remove(wb.active)
        ws1 = wb.create_sheet("Параметры")
        ws1.append(["Параметр", "Значение"])
        ws1.append(["Продукт", result['product_name']])
        ws1.append(["Количество", result['Q']])
        ws1.append(["Размер наряда", result['N']])
        ws1.append(["Календарное время (ч)", result['T']])
        ws1.append(["Рабочих дней", result['days_needed']])
        ws1.append(["Трудоёмкость (чел·ч)", result['total_labor']])
        if result['is_glue']:
            ws1.append(["Общий вес (г)", result['total_weight']])
            ws1.append(["Необходимо 4-кг канистр", result['can_count_4kg']])
            ws1.append(["Недостаток в последней 4-кг канистре (г)", result['shortage_4kg']])
            ws1.append(["Необходимо 1-кг канистр", result['can_count_1kg']])
            ws1.append(["Недостаток в последней 1-кг канистре (г)", result['shortage_1kg']])
            if result['corrected']:
                ws1.append(["Корректировка", "Выполнена"])
        ws2 = wb.create_sheet("Операции")
        ws2.append(["Операция", "t_i (ч)", "Наладка (ч)", "Людей", "Общее время (ч)", "Дней работы"])
        for i, name in enumerate(result['name_list']):
            ws2.append([name, result['t_list'][i], result['setup_list'][i],
                       result['people_list'][i], result['m'] * result['t_list'][i],
                       result['days_work_list'][i]])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        st.download_button("📥 Скачать Excel-отчёт", buffer,
                           file_name=f"report_{result['product_name'].replace(' ', '_')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.error(f"Ошибка Excel: {e}")
